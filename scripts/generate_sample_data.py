"""Generate Kafka-ready sales transactions from the source workbook.

Without ``--target-records`` the script preserves the original behaviour and
creates one transaction for every valid workbook row. With
``--target-records`` it creates a deterministic synthetic transaction-level
dataset calibrated from the real products, prices, costs, quantities, and
return rate in the workbook.
"""

import argparse
import calendar
import json
import os
import random
from collections import defaultdict
from datetime import datetime, timezone

import openpyxl


SHEET_TO_MONTH = {
    "Nov_file": (2025, 11),
    "Dec_file": (2025, 12),
    "Jan_file": (2026, 1),
}

# Transparent retail-seasonality assumptions for synthetic event volume.
MONTH_VOLUME_WEIGHTS = {
    "Nov_file": 0.30,
    "Dec_file": 0.42,
    "Jan_file": 0.28,
}
HOUR_WEIGHTS = {
    9: 2,
    10: 4,
    11: 6,
    12: 8,
    13: 8,
    14: 6,
    15: 5,
    16: 6,
    17: 8,
    18: 10,
    19: 8,
    20: 4,
}
WEEKDAY_WEIGHTS = {
    0: 1.0,
    1: 1.0,
    2: 1.0,
    3: 1.05,
    4: 1.20,
    5: 1.35,
    6: 1.15,
}

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_XLSX = os.path.join(BASE_DIR, "data", "Master_Sales_data.xlsx")
OUT_PATH = os.path.join(BASE_DIR, "sample_data", "sales_transactions.json")


def _load_source_rows(source_path=SOURCE_XLSX):
    """Return valid source rows grouped by workbook sheet."""
    workbook = openpyxl.load_workbook(source_path, data_only=True)
    rows_by_sheet = defaultdict(list)

    for sheet_name in SHEET_TO_MONTH:
        worksheet = workbook[sheet_name]
        headers = [
            cell.value
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
        columns = {header: index for index, header in enumerate(headers)}

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            stock_code = row[columns["Stock Code"]]
            quantity = row[columns["Sold Period"]]
            unit_cost = row[columns["Unit Cost"]]
            unit_price = row[columns["Unit Price"]]
            sales_value = row[columns["Sales Value"]]
            if any(
                value is None
                for value in (stock_code, quantity, unit_cost, unit_price)
            ):
                continue

            rows_by_sheet[sheet_name].append(
                {
                    "product_id": str(stock_code),
                    "quantity": int(quantity),
                    "unit_cost": round(float(unit_cost), 2),
                    "price": round(float(unit_price), 2),
                    "sales_value": (
                        round(float(sales_value), 2)
                        if sales_value is not None
                        else None
                    ),
                }
            )

    return rows_by_sheet


def _allocate_month_counts(target_records):
    """Allocate an exact target across months using largest remainders."""
    exact = {
        sheet: target_records * weight
        for sheet, weight in MONTH_VOLUME_WEIGHTS.items()
    }
    counts = {sheet: int(value) for sheet, value in exact.items()}
    remaining = target_records - sum(counts.values())
    by_remainder = sorted(
        exact,
        key=lambda sheet: exact[sheet] - counts[sheet],
        reverse=True,
    )
    for sheet in by_remainder[:remaining]:
        counts[sheet] += 1
    return counts


def _random_timestamp(rng, year, month):
    days = list(range(1, calendar.monthrange(year, month)[1] + 1))
    day_weights = [
        WEEKDAY_WEIGHTS[datetime(year, month, day).weekday()]
        for day in days
    ]
    day = rng.choices(days, weights=day_weights, k=1)[0]
    hour = rng.choices(
        list(HOUR_WEIGHTS),
        weights=list(HOUR_WEIGHTS.values()),
        k=1,
    )[0]
    return datetime(
        year,
        month,
        day,
        hour,
        rng.randrange(60),
        rng.randrange(60),
        tzinfo=timezone.utc,
    )


def _build_original_transactions(rows_by_sheet):
    transactions = []
    tx_counter = 0
    for sheet_name, (year, month) in SHEET_TO_MONTH.items():
        days_in_month = calendar.monthrange(year, month)[1]
        for month_counter, source in enumerate(
            rows_by_sheet[sheet_name],
            start=1,
        ):
            day = ((month_counter - 1) % days_in_month) + 1
            tx_counter += 1
            quantity = source["quantity"]
            transactions.append(
                {
                    "transaction_id": (
                        f"TXN-{sheet_name[:3].upper()}-{tx_counter:06d}"
                    ),
                    "timestamp": (
                        f"{year:04d}-{month:02d}-{day:02d}T00:00:00Z"
                    ),
                    "store_id": "BELFAST",
                    "product_id": source["product_id"],
                    "quantity": quantity,
                    "price": source["price"],
                    "unit_cost": source["unit_cost"],
                    "cost_sales": round(source["unit_cost"] * quantity, 2),
                    "sales_value": source["sales_value"],
                }
            )
    return transactions


def _build_synthetic_transactions(rows_by_sheet, target_records, seed):
    if target_records <= 0:
        raise ValueError("target_records must be greater than zero")

    rng = random.Random(seed)
    month_counts = _allocate_month_counts(target_records)
    all_quantities = [
        row["quantity"]
        for rows in rows_by_sheet.values()
        for row in rows
    ]
    positive_quantities = [
        min(quantity, 12)
        for quantity in all_quantities
        if quantity > 0
    ]
    return_quantities = [
        max(quantity, -12)
        for quantity in all_quantities
        if quantity < 0
    ]
    return_rate = len(return_quantities) / len(all_quantities)

    transactions = []
    tx_counter = 0
    for sheet_name, (year, month) in SHEET_TO_MONTH.items():
        templates = rows_by_sheet[sheet_name]
        product_weights = [
            max(abs(row["quantity"]), 1)
            for row in templates
        ]

        for _ in range(month_counts[sheet_name]):
            source = rng.choices(
                templates,
                weights=product_weights,
                k=1,
            )[0]
            is_return = bool(return_quantities) and rng.random() < return_rate
            if is_return:
                quantity = rng.choice(return_quantities)
            else:
                quantity = rng.choice(positive_quantities)

            timestamp = _random_timestamp(rng, year, month)
            tx_counter += 1
            transactions.append(
                {
                    "transaction_id": (
                        f"TXN-SYN-{year}{month:02d}-{tx_counter:06d}"
                    ),
                    "timestamp": timestamp.isoformat().replace("+00:00", "Z"),
                    "store_id": "BELFAST",
                    "product_id": source["product_id"],
                    "quantity": quantity,
                    "price": source["price"],
                    "unit_cost": source["unit_cost"],
                    "cost_sales": round(source["unit_cost"] * quantity, 2),
                    "sales_value": round(source["price"] * quantity, 2),
                }
            )

    # Kafka receives the generated events in event-time order.
    transactions.sort(key=lambda transaction: transaction["timestamp"])
    return transactions


def build_transactions(
    target_records=None,
    seed=42,
    source_path=SOURCE_XLSX,
    out_path=OUT_PATH,
):
    rows_by_sheet = _load_source_rows(source_path)
    if target_records is None:
        transactions = _build_original_transactions(rows_by_sheet)
        dataset_type = "source-derived"
    else:
        transactions = _build_synthetic_transactions(
            rows_by_sheet,
            target_records,
            seed,
        )
        dataset_type = f"synthetic (seed={seed})"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as output_file:
        json.dump(transactions, output_file, indent=2)

    print(
        f"Wrote {len(transactions)} {dataset_type} transactions to {out_path}"
    )
    if target_records is not None:
        counts = _allocate_month_counts(target_records)
        allocation = ", ".join(
            f"{sheet}={count}"
            for sheet, count in counts.items()
        )
        print(f"Monthly allocation: {allocation}")
    return transactions


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate RetailBatch sample transactions"
    )
    parser.add_argument(
        "--target-records",
        type=int,
        default=None,
        help=(
            "Generate this many realistic synthetic transactions; "
            "omit for the original 966 rows"
        ),
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="Random seed for reproducible generation (default: 42)",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    build_transactions(
        target_records=args.target_records,
        seed=args.seed,
    )
